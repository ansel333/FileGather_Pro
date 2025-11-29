import re
from pathlib import Path

def exact_match_filename(filename, keyword):
    """
    精确查找：文件名必须严格对应关键词
    不考虑扩展名，只匹配文件名主体
    """
    if not keyword:
        return True
    
    # 获取文件名（不包括扩展名）
    name_without_ext = Path(filename).stem.lower()
    keyword_lower = keyword.strip().lower()
    
    # 精确匹配：文件名必须完全等于关键词
    return name_without_ext == keyword_lower

def matches_keyword(text, keyword):
    if not keyword:
        return True

    exact_matches = re.findall(r'"([^"]*)"', keyword)
    for exact in exact_matches:
        if exact.lower() not in text.lower():
            return False
        keyword = keyword.replace(f'"{exact}"', '')

    tokens = re.split(r'[\s\n]+', keyword.strip())
    must_include = []
    must_exclude = []
    any_include = []

    for token in tokens:
        if not token:
            continue
        if token.startswith('+'):
            must_include.append(token[1:].lower())
        elif token.startswith('-'):
            must_exclude.append(token[1:].lower())
        elif '|' in token:
            any_include.extend([t.lower() for t in token.split('|')])
        else:
            any_include.append(token.lower())

    for term in must_include:
        if term not in text.lower():
            return False

    for term in must_exclude:
        if term in text.lower():
            return False

    if any_include:
        found = False
        for term in any_include:
            if '*' in term:
                # Simple wildcard matching
                if re.search(term.replace('*', '.*'), text.lower()):
                    found = True
                    break
            elif term in text.lower():
                found = True
                break
        if not found:
            return False

    return True

def search_content(file_path, keyword):
    try:
        file_path = Path(file_path)
        ext = file_path.suffix.lower()

        if ext in ['.txt', '.py', '.java', '.cpp', '.h', '.html', '.css', '.js', '.csv', '.ini', '.log']:
            return search_text_file(file_path, keyword)
        elif ext == '.pdf':
            return search_pdf(file_path, keyword)
        elif ext in ['.docx']:
            return search_docx(file_path, keyword)
        elif ext in ['.xlsx']:
            return search_excel(file_path, keyword)
        else:
            return False
    except Exception as e:
        print(f"内容搜索失败: {file_path} - {str(e)}")
        return False

def search_text_file(file_path, keyword):
    encodings = ['utf-8', 'gbk', 'latin-1']
    for encoding in encodings:
        try:
            with file_path.open('r', encoding=encoding, errors='ignore') as f:
                content = f.read(3000)
                return matches_keyword(content, keyword)
        except UnicodeDecodeError:
            continue
    return False

def search_pdf(file_path, keyword):
    try:
        from fitz import fitz  # 使用PyMuPDF
        doc = fitz.open(str(file_path))
        text = ""
        for page in doc:
            text += page.get_text()
            if len(text) > 3000:
                break
        doc.close()
        return matches_keyword(text, keyword)
    except ImportError:
        return search_text_file(file_path, keyword)
    except Exception:
        return False

def search_docx(file_path, keyword):
    try:
        from docx import Document
        doc = Document(str(file_path))
        text = ""
        for para in doc.paragraphs:
            text += para.text + " "
            if len(text) > 3000:
                break
        return matches_keyword(text, keyword)
    except ImportError:
        return search_text_file(file_path, keyword)
    except Exception:
        return False

def search_excel(file_path, keyword):
    try:
        from openpyxl import load_workbook
        wb = load_workbook(str(file_path), read_only=True)
        text = ""
        for sheet in wb:
            for row in sheet.iter_rows(values_only=True):
                for cell in row:
                    if cell:
                        text += str(cell) + " "
                if len(text) > 3000:
                    break
            if len(text) > 3000:
                break
        wb.close()
        return matches_keyword(text, keyword)
    except ImportError:
        return search_text_file(file_path, keyword)
    except Exception:
        return False
